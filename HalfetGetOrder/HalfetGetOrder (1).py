# 1. 라이브러리 불러오기
import os
import time
import hmac, hashlib
import urllib.parse
import urllib.request
import ssl
import requests
import openpyxl
import json
import unicodedata
import math
from datetime import date, datetime, timedelta
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import xmltodict
# from keys import partner_key, godo_key, cp_accesskey, cp_secretkey

# ─────────────────────────────────────────────────────────
# 공통 사용 변수
# ─────────────────────────────────────────────────────────
content_type = "application/json;charset=UTF-8"

os.environ['TZ'] = 'GMT+0'
datetime_signed = time.strftime('%y%m%d') + 'T' + time.strftime('%H%M%S') + 'Z'
method = "GET"

# 날짜범위 - 시작날짜(오늘날짜 - 7일) / 종료날짜(오늘날짜)
created_from = str(date.today() - timedelta(days=7))
created_to   = str(date.today())
print("조회기간:", created_from, "~", created_to)

# ─────────────────────────────────────────────────────────
# API 키/기본 설정
# ─────────────────────────────────────────────────────────
# 쿠팡 API KEY
cp_accesskey = "7b6058d9-7745-4cf8-9881-cf4a469c0512"
cp_secretkey = "a4e91ae56e91a47b696712dd29008a405e8d4c25"

# 고도몰 API KEY
partner_key = "MCVFRCVENSVCNiUyOCUxMSVEQ28="
godo_key = "JTBBJUNFJTQwZCUxMiVCOCUzQSVBRCVDQW0lRjMlODclRjUlQUElQjIlQ0QlRUYlMDMlQkUlQzklREIlOTh6bU0lMTElNUIlREElM0I0JTFCJUZCcCUxMyVFRCU0MCU3RTUlMUYlRjk="

# ─────────────────────────────────────────────────────────
# 추가옵션 캐시 설정
# ─────────────────────────────────────────────────────────
ADD_GOODS_CACHE = "godo_add_goods_cache.json"
REFRESH_ADD_GOODS = False  # 새 옵션 갱신 시 True로 1회 실행

# ─────────────────────────────────────────────────────────
# 경로(엑셀 저장 위치)
# ─────────────────────────────────────────────────────────
BASE_DIR = r"C:/Users/UserK/Desktop/[작업물]/[판매] 주문수집 프로그램"
os.makedirs(BASE_DIR, exist_ok=True)
today_date = date.today().strftime('%Y%m%d')
WAYBILL_XLSX = os.path.join(BASE_DIR, f"대한통운 송장등록_{today_date}.xlsx")
ORDER_XLSX   = os.path.join(BASE_DIR, f"주문수집_{today_date}.xlsx")

# ─────────────────────────────────────────────────────────
# 공통 유틸/헬퍼
# ─────────────────────────────────────────────────────────
def first_non_empty(*vals):
    for v in vals:
        if v is not None and str(v).strip() != "":
            return v
    return ""

def _to_int(x, default=0):
    try:
        return int(float(str(x).strip()))
    except Exception:
        return default

def _to_float(x, default=0.0):
    try:
        return float(str(x).strip())
    except Exception:
        return default

def _as_list(x):
    if x is None:
        return []
    return x if isinstance(x, list) else [x]

def _fmt_dt(s):
    """문자열 날짜를 'YYYY.MM.DD HH:MM'로 최대한 정규화"""
    if not s:
        return ""
    s = str(s).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y.%m.%d %H:%M")
        except Exception:
            pass
    # yyyy-mm-dd만 온 경우
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%Y.%m.%d 00:00")
    except Exception:
        return s

def get_box_count_from_items(items):
    """3대까지 1박스, 이후 4대마다 1박스 추가"""
    if not items:
        return 1
    try:
        total_qty = sum(int(i.get('quantity', 1)) for i in items)
        if total_qty <= 3:
            return 1
        return 1 + math.ceil((total_qty - 3) / 4)
    except Exception:
        return 1

def visual_len(s):
    """전각(W/F/A)은 2, 그 외는 1로 계산해서 가시 길이 추정"""
    if s is None:
        return 0
    total = 0
    for ch in str(s):
        ea = unicodedata.east_asian_width(ch)
        total += 2 if ea in ('W', 'F', 'A') else 1
    return total


# ─────────────────────────────────────────────────────────
# 추가옵션 캐시 로드/저장
# ─────────────────────────────────────────────────────────
def load_add_goods_map_from_cache(path=ADD_GOODS_CACHE):
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return {str(k): {"goodsNm": v.get("goodsNm",""),
                                 "goodsPrice": float(v.get("goodsPrice",0.0))}
                        for k, v in data.items()}
    except Exception as e:
        print("추가옵션 캐시 로드 오류:", e)
    return {}

def save_add_goods_map_to_cache(add_map, path=ADD_GOODS_CACHE):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(add_map, f, ensure_ascii=False, indent=2)
        print(f"추가옵션 캐시 저장 완료: {path} (항목 {len(add_map)}개)")
    except Exception as e:
        print("추가옵션 캐시 저장 오류:", e)

# ─────────────────────────────────────────────────────────
# 고도몰 추가상품 목록 조회 (캐시 우선)
# ─────────────────────────────────────────────────────────
def fetch_godomall_add_goods_map():
    add_goods_map = load_add_goods_map_from_cache()
    if REFRESH_ADD_GOODS or not add_goods_map:
        print("추가옵션 캐시가 없거나 갱신 플래그가 True → OpenHub에서 새로 수집합니다.")
        godo_add_url = (
            "https://openhub.godo.co.kr/godomall5/goods/Goods_Add_Search.php?"
            f"partner_key={partner_key}&key={godo_key}"
        )
        r = requests.post(godo_add_url, timeout=30)
        ctype2 = (r.headers.get('Content-Type') or '').lower()
        if 'euc-kr' in ctype2 or 'cp949' in ctype2:
            r.encoding = 'cp949'
        elif not r.encoding:
            r.encoding = 'utf-8'
        add_xml = r.text
        add_json = xmltodict.parse(add_xml)

        def _txt(d, k):
            v = d.get(k)
            return v.strip() if isinstance(v, str) else ("" if v is None else str(v))
        def _price(d):
            for k in ("goodsPrice", "price", "salePrice", "fixedPrice", "consumerPrice"):
                if d.get(k) not in (None, ""):
                    try:
                        return float(str(d[k]).strip())
                    except:
                        pass
            return 0.0

        root_add = add_json.get("data", {}) if isinstance(add_json, dict) else {}
        ret_add  = root_add.get("return", {}) or {}
        cands = None
        for k in ("add_goods_data", "addGoodsData", "goods_data", "goodsData", "list", "rows"):
            if ret_add.get(k) is not None:
                cands = ret_add.get(k)
                break

        items = _as_list(cands)
        if items and isinstance(items[0], dict) and any(subk in items[0] for subk in ("item", "goods", "row")):
            tmp = []
            for it in items:
                for subk in ("item", "goods", "row"):
                    if it.get(subk) is not None:
                        tmp.extend(_as_list(it[subk]))
            if tmp:
                items = tmp

        add_goods_map = {}
        for it in items:
            add_no = _txt(it, "addGoodsNo") or _txt(it, "add_goods_no") or _txt(it, "goodsNo")
            name   = _txt(it, "goodsNm") or _txt(it, "goodsNmStandard") or _txt(it, "goods_name")
            price  = _price(it)
            if add_no and name:
                add_goods_map[add_no] = {"goodsNm": name, "goodsPrice": price}

        print("추가상품 정의 수(신규 수집):", len(add_goods_map))
        save_add_goods_map_to_cache(add_goods_map)
    else:
        print("추가옵션 캐시 사용:", len(add_goods_map), "개")
    return add_goods_map

add_goods_map = fetch_godomall_add_goods_map()

# ─────────────────────────────────────────────────────────
# 고도몰 주문 조회 (XML)
# ─────────────────────────────────────────────────────────
godo_order_url = (
    "https://openhub.godo.co.kr/godomall5/order/Order_Search.php?"
    f"partner_key={partner_key}&key={godo_key}"
    f"&startDate={created_from}&endDate={created_to}"
    "&dateType=order&orderStatus=g1"
)
godo_req = requests.post(godo_order_url, timeout=30)
ctype = (godo_req.headers.get('Content-Type') or '').lower()
if 'euc-kr' in ctype or 'cp949' in ctype:
    godo_req.encoding = 'cp949'
elif not godo_req.encoding:
    godo_req.encoding = 'utf-8'
godo_xml_text = godo_req.text
godo_json = xmltodict.parse(godo_xml_text)

# ─────────────────────────────────────────────────────────
# 쿠팡 api 연동
# ─────────────────────────────────────────────────────────
cp_domain = "https://api-gateway.coupang.com"
cp_venderId = "A01093941"

cp_path = f"/v2/providers/openapi/apis/api/v4/vendors/{cp_venderId}/ordersheets"
cp_query = urllib.parse.urlencode({
    "createdAtFrom": created_from,
    "createdAtTo": created_to,
    "status": "INSTRUCT"
})
message = datetime_signed + method + cp_path + cp_query
signature = hmac.new(cp_secretkey.encode('utf-8'), message.encode('utf-8'), hashlib.sha256).hexdigest()
authorization = (
    f"CEA algorithm=HmacSHA256, access-key={cp_accesskey}, signed-date={datetime_signed}, signature={signature}"
)
cp_url = f"{cp_domain}{cp_path}?{cp_query}"
req = urllib.request.Request(cp_url)
req.add_header("Content-type", content_type)
req.add_header("Authorization", authorization)
req.get_method = lambda: method

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

try:
    resp = urllib.request.urlopen(req, context=ctx)
    coupang_body = resp.read().decode(resp.headers.get_content_charset() or "utf-8")
except Exception as e:
    print("쿠팡 API 호출 오류:", e)
    coupang_body = ""

# ─────────────────────────────────────────────────────────
# 정규화: 쿠팡 / 고도몰
# ─────────────────────────────────────────────────────────
def normalize_coupang_orders(coupang_body):
    try:
        data = json.loads(coupang_body)
    except Exception:
        return []
    orders = data.get('data') or data.get('content') or []
    norm = []
    for od in orders:
        ship = od.get('shippingAddress') or {}
        recv = od.get('receiver') or {}
        orderer = od.get('orderer', {}) or {}
        name = first_non_empty(ship.get('name'), recv.get('name'), orderer.get('name'))
        phone = first_non_empty(
            ship.get('safeNumber'), recv.get('safeNumber'),
            ship.get('phone'), ship.get('phoneNo'),
            recv.get('receiverPhone'), orderer.get('phone')
        )
        addr1 = first_non_empty(ship.get('address1'), recv.get('addr1'))
        addr2 = first_non_empty(ship.get('address2'), recv.get('addr2'))
        zipcode = first_non_empty(ship.get('zipcode'), recv.get('zipCode'))
        items_raw = od.get('orderItems', []) or []
        items = [{"quantity": _to_int(first_non_empty(it.get('shippingCount'), it.get('quantity'), 1), 1)}
                 for it in items_raw]
        norm.append({
            "channel": "coupang",
            "name": name, "phone": phone, "addr1": addr1, "addr2": addr2,
            "zipcode": zipcode, "items": items,
            "raw": od
        })
    return norm

def normalize_godomall_orders_from_xml(json_data):
    try:
        root = json_data.get('data', {})
        ret = root.get('return', {}) or {}
        od_list = _as_list(ret.get('order_data'))
    except Exception:
        return []

    norm = []
    for od in od_list:
        info = od.get('orderInfoData') or {}
        def get_text(d, key):
            v = d.get(key)
            return (v or "").strip() if isinstance(v, str) else (v or "")
        name = get_text(info, 'receiverName')
        safe_fl = (get_text(info, 'receiverUseSafeNumberFl') or '').lower() == 'y'
        safe_no = get_text(info, 'receiverSafeNumber')
        phone = safe_no if (safe_fl and safe_no) else (get_text(info, 'receiverPhone') or get_text(info, 'receiverCellPhone'))
        zipcode = get_text(info, 'receiverZonecode') or get_text(info, 'receiverZipcode')
        addr1 = get_text(info, 'receiverAddress')
        addr2 = get_text(info, 'receiverAddressSub')
        og_list = _as_list(od.get('orderGoodsData'))
        if not addr1 and og_list:
            va = og_list[0].get('visitAddress')
            if isinstance(va, str) and va.strip():
                addr1 = va.strip()
                addr2 = ""
        items = []
        for og in og_list:
            qty_raw = og.get('goodsCnt', 1)
            try:
                qty = int(str(qty_raw).strip() or '1')
            except Exception:
                qty = 1
            items.append({'quantity': qty})
        norm.append({
            "channel": "godomall",
            "name": name, "phone": phone, "addr1": addr1,
            "addr2": addr2, "zipcode": zipcode, "items": items,
            "raw": od
        })
    return norm

# ─────────────────────────────────────────────────────────
# ① 대한통운 송장등록용 엑셀 생성
#   - 요청사항: 고도몰 주문 제외(쿠팡만 기록)
#   - 배송메세지2에 플랫폼 정보(쿠팡) 기입
# ─────────────────────────────────────────────────────────
wb1 = openpyxl.Workbook()
sheet1 = wb1.active
sheet1.title = "판매 주문수집"

first_col1 = [
    '예약구분', '집하예정일', '받는분성명', '받는분전화번호', '받는분기타연락처',
    '받는분우편번호', '받는분주소(전체, 분할)', '운송장번호', '고객주문번호',
    '품목명', '박스수량', '박스타입', '기본운임', '배송메세지1',
    '배송메세지2', '품목명', '운임구분'
]
sheet1.append(first_col1)

header_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
header_font = Font(bold=True)
header_align = Alignment(horizontal="center", vertical="center")
for cell in sheet1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# 쿠팡만
orders_coupang_only = []
if coupang_body:
    orders_coupang_only += normalize_coupang_orders(coupang_body)

for od in orders_coupang_only:
    name = od.get("name", "")
    phone = od.get("phone", "")
    addr1 = od.get("addr1", "")
    addr2 = od.get("addr2", "")
    zipcode = od.get("zipcode", "")
    address = f"{addr1} {addr2}".strip()
    box_cnt = get_box_count_from_items(od.get("items", []))

    # 배송메세지2 = 플랫폼
    platform_name = "쿠팡"

    row = [
        "일반",
        date.today().strftime('%Y%m%d'),
        name, phone, "", zipcode, address,
        "", "", "", box_cnt, "", "", "",
        platform_name, "", ""
    ]
    sheet1.append(row)

# 가운데 정렬 및 열 너비 자동 조정
center_align = Alignment(horizontal="center", vertical="center")
for row in sheet1.iter_rows():
    for cell in row:
        cell.alignment = center_align

for column_cells in sheet1.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in column_cells)
    col_letter = column_cells[0].column_letter
    sheet1.column_dimensions[col_letter].width = max_len * 1.3 + 2

# 저장 (덮어쓰기)
if os.path.exists(WAYBILL_XLSX):
    try: os.remove(WAYBILL_XLSX)
    except: pass
wb1.save(WAYBILL_XLSX)
print(f"✅ 엑셀 저장 완료: {WAYBILL_XLSX}")

# ─────────────────────────────────────────────────────────
# ② 주문수집(쿠팡 + 고도몰) 엑셀 생성
#   - 요구사항 반영:
#     (1) 주문별 블록 테두리 (H열 포함)
#     (2) 수취인 이름 하위행 병합
#     (3) 상품 결제 금액 뒤에 '원' 붙여 표기
#     (4) 고도몰: 본상품=부모행, 추가옵션=하위행
#     (5) 등록옵션명(고도몰): goodsCd
#     (6) 기본 행 높이(25~26) — ‘상품명+옵션명’ 줄바꿈 행 제외하고 적용
#     (7) 블록 마지막 행 두꺼운 하단 테두리(병합된 수취인 이름 셀 포함)
# ─────────────────────────────────────────────────────────
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = "주문내역"

# 헤더 행 (플랫폼 포함)
header_cols = ['플랫폼', '주문일시', '총 상품결제금액', '수취인 이름', '상품명 + 옵션명', '수량', '수취인 전화번호', '등록옵션명']
sheet2.append(header_cols)

# 스타일
for cell in sheet2[1]:
    cell.fill = header_fill
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style="thin", color="000000")
thick = Side(style="thick", color="000000")

# ── 추가: 블록 테두리(H열 포함, end_col ‘포함’) 및 굵은 하단선 유틸
def apply_border_block(ws, start_row, end_row, start_col=1, end_col=8):
    """주문 블록에 얇은 테두리 적용 (H열 포함, end_col 포함)"""
    for r in range(start_row, end_row+1):
        for c in range(start_col, end_col+1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_thick_bottom(ws, block_start, block_end, start_col=1, end_col=8):
    """블록의 마지막 행에 두꺼운 하단 테두리 적용 + 병합된 수취인 이름셀 하단선 보강"""
    # 마지막 행 전체(1..8) 두꺼운 하단선
    for c in range(start_col, end_col+1):
        cell = ws.cell(row=block_end, column=c)
        cell.border = Border(
            left=cell.border.left or thin,
            right=cell.border.right or thin,
            top=cell.border.top or thin,
            bottom=thick
        )
    # 병합된 수취인 이름셀(컬럼 4)의 하단도 시각적으로 확실히 보이도록 보강
    # (openpyxl 병합셀 스타일 특성 대응: 상단왼쪽 셀에도 한 번 더 하단 두께를 강조)
    top_left = ws.cell(row=block_start, column=4)
    top_left.border = Border(
        left=top_left.border.left or thin,
        right=top_left.border.right or thin,
        top=top_left.border.top or thin,
        bottom=thick
    )

def merge_receiver_name(ws, start_row, end_row):
    """수취인 이름 열(D열, 4번)을 블록 단위로 병합"""
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
        ws.cell(row=start_row, column=4).alignment = center

current_row = 2  # 데이터 시작 행

# ========== 2-1) 쿠팡 → 블록 ==========
platform_name = "쿠팡"
if coupang_body:
    try:
        resp_json = json.loads(coupang_body)
        orders = resp_json.get('data') or resp_json.get('content', []) or []

        for od in orders:
            block_start = current_row

            ordered_at = _fmt_dt(od.get('orderedAt') or od.get('orderDate', ''))

            # 상품 결제금액 합계 ('원' 붙임)
            total_price = 0.0
            for item in od.get('orderItems', []):
                price = _to_float(item.get('orderPrice', item.get('price', 0)))
                qty = _to_int(item.get('shippingCount', 1), 1)
                total_price += price * qty
            total_price_str = f"{_to_int(total_price):,}원"

            receiver_name = (
                (od.get('shippingAddress') or {}).get('name', '') or
                (od.get('receiver') or {}).get('name', '')
            )

            # 상품명 + 옵션명 (쿠팡: 한 줄 표기)
            item_names = []
            total_qty = 0
            for item in od.get('orderItems', []):
                name = item.get('sellerProductName') or item.get('vendorItemName') or item.get('productName') or ""
                option = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
                qty = _to_int(item.get('shippingCount', 1), 1)
                total_qty += qty
                if name and option and option != name:
                    item_names.append(f"{name} / {option}")
                else:
                    item_names.append(name or option)
            product_info = " / ".join([x for x in item_names if x])
            total_qty = total_qty or 1

            phone = (
                (od.get('shippingAddress') or {}).get('safeNumber') or
                (od.get('receiver') or {}).get('safeNumber') or
                (od.get('receiver') or {}).get('phone') or
                (od.get('receiver') or {}).get('receiverPhone') or ''
            )

            # 등록옵션명(쿠팡: 옵션명 모음)
            option_names = []
            for item in od.get('orderItems', []):
                option_name = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
                if option_name:
                    option_names.append(str(option_name))
            option_name_str = ", ".join(option_names)

            # 부모행 1줄만
            sheet2.append([
                platform_name, ordered_at, total_price_str, receiver_name,
                product_info, total_qty, phone, option_name_str
            ])
            current_row += 1

            # 블록 테두리 + 수취인 이름 병합(쿠팡은 1줄)
            block_end = current_row - 1
            apply_border_block(sheet2, block_start, block_end, 1, 8)  # H열 포함
            merge_receiver_name(sheet2, block_start, block_end)
            apply_thick_bottom(sheet2, block_start, block_end, 1, 8)   # 마지막 행 두꺼운 하단선

    except Exception as e:
        print("⚠️ 쿠팡 JSON 파싱 오류:", e)


# ========== 2-2) 고도몰 → 세트 표시(본상품 + 그 본상품의 추가옵션들: parentGoodsNo 매칭) ==========
platform_name = "고도몰"

try:
    root = godo_json.get('data', {}) if isinstance(godo_json, dict) else {}
    ret = root.get('return', {}) or {}
    od_list = _as_list(ret.get('order_data'))

    for od in od_list:
        block_start = current_row

        ordered_at = _fmt_dt(od.get('orderDate', ''))
        info = od.get('orderInfoData') or {}
        receiver_name = (info.get('receiverName') or "").strip()
        safe_fl = str(info.get('receiverUseSafeNumberFl') or '').strip().lower() == 'y'
        safe_no = (info.get('receiverSafeNumber') or '').strip()
        phone = safe_no if (safe_fl and safe_no) else (
            (info.get('receiverPhone') or info.get('receiverCellPhone') or '').strip()
        )

        parents = _as_list(od.get('orderGoodsData'))
        adds    = _as_list(od.get('addGoodsData') or od.get('orderAddGoodsData'))

        # Map parents by goodsNo (string)
        parent_index = {}
        for idx, og in enumerate(parents):
            gno = str(og.get('goodsNo') or "").strip()
            if gno:
                parent_index[gno] = idx

        # Group add options by parentGoodsNo (strict)
        adds_by_parent_idx = {i: [] for i in range(len(parents))}
        for add in adds:
            pno = str(add.get('parentGoodsNo') or "").strip()
            if pno and pno in parent_index:
                adds_by_parent_idx[parent_index[pno]].append(add)
            else:
                # 요청: 미지정 옵션은 출력하지 않음
                pass

        # Output sets: [parent] then its add options
        first_parent = True
        for i, og in enumerate(parents):
            goodsCd  = (og.get('goodsCd') or '').strip()
            goodsNm  = (og.get('goodsNm') or og.get('goodsNmStandard') or '').strip()
            opt_text = (og.get('optionTextInfo') or '').strip()
            qty      = _to_int(og.get('goodsCnt', 1), 1)
            price    = _to_float(og.get('goodsPrice', 0.0), 0.0)

            # total = parent + its adds
            total_price = price * (qty or 1)
            for add in adds_by_parent_idx.get(i, []):
                add_qty  = _to_int(add.get('goodsCnt', 1), 1)
                add_price = _to_float(add.get('goodsPrice', 0.0), 0.0)
                total_price += add_price * add_qty

            total_price_str = f"{_to_int(total_price):,}원"
            product_info_parent = f"{goodsCd} / {opt_text}" if opt_text else (goodsNm or goodsCd)
            reg_option_value = goodsCd

            sheet2.append([
                platform_name,
                ordered_at if first_parent else "",
                total_price_str,
                receiver_name if first_parent else "",
                product_info_parent,
                (qty or 1),
                phone if first_parent else "",
                reg_option_value
            ])
            current_row += 1
            first_parent = False

            for add in adds_by_parent_idx.get(i, []):
                add_name = (add.get('goodsNm') or add.get('goodsNmStandard') or "").strip()
                add_qty  = _to_int(add.get('goodsCnt', 1), 1)
                sheet2.append(["", "", "", "", f"+ {add_name}", add_qty, "", ""])
                current_row += 1

        block_end = current_row - 1
        if block_end >= block_start:
            apply_border_block(sheet2, block_start, block_end, 1, 8)
            merge_receiver_name(sheet2, block_start, block_end)
            apply_thick_bottom(sheet2, block_start, block_end, 1, 8)

except Exception as e:
    print("⚠️ 고도몰 XML 파싱 오류(parentGoodsNo 매칭):", e)
# 2-마지막) 시트 서식(가운데 정렬 + 열 너비 자동 + 행 높이)
min_widths = {
    '플랫폼': 8,
    '주문일시': 16,
    '상품결제금액': 14,
    '수취인 이름': 20,
    '상품명 + 옵션명': 32,
    '수량': 10,
    '수취인 전화번호': 16,
    '등록옵션명': 50,
}
headers = [cell.value for cell in sheet2[1]]
for col in sheet2.columns:
    col_idx = col[0].column
    col_letter = get_column_letter(col_idx)
    header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ''

    max_len = visual_len(header)
    for cell in col:
        vlen = visual_len(cell.value)
        if vlen > max_len:
            max_len = vlen
        # 기본: 가운데 정렬
        if header == '상품명 + 옵션명' and vlen > 40:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        # 등록옵션명: 텍스트 서식
        if header == '등록옵션명':
            cell.number_format = '@'

    auto_width = int(max_len * 0.5)
    if header == '등록옵션명':
        auto_width = int(max_len * 0.5) + 4
    target_width = max(auto_width, min_widths.get(header, 12))
    sheet2.column_dimensions[col_letter].width = target_width

# ── 추가: 기본 행 높이 26 적용 (단, '상품명 + 옵션명' 셀 wrap_text=True 인 행은 제외)
#      ※ 쿠팡/고도몰 구분 없이 시트 전체 데이터 행에 적용
for r in range(2, sheet2.max_row + 1):
    prod_cell = sheet2.cell(row=r, column=5)  # '상품명 + 옵션명'
    pclen = visual_len(prod_cell.value)
    if pclen > 40:
        sheet2.row_dimensions[r].height = 34
    if not (prod_cell.alignment and prod_cell.alignment.wrap_text):
        sheet2.row_dimensions[r].height = 24

# 저장(덮어쓰기)
if os.path.exists(ORDER_XLSX):
    try: os.remove(ORDER_XLSX)
    except: pass
wb2.save(ORDER_XLSX)
print(f"✅ 엑셀 저장 완료: {ORDER_XLSX}")
