
import os
import openpyxl
from datetime import date
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from .utils import visual_len, _to_int, _to_float
from .utils import _fmt_dt

header_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
thin = Side(style="thin", color="000000")
thick = Side(style="thick", color="000000")

def create_orders_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주문내역"
    headers = ['플랫폼', '주문일시', '총 상품결제금액', '수취인 이름', '상품명 + 옵션명', '수량', '수취인 전화번호', '등록옵션명']
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
    return wb, ws

def apply_border_block(ws, start_row, end_row, start_col=1, end_col=8):
    for r in range(start_row, end_row+1):
        for c in range(start_col, end_col+1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_thick_bottom(ws, block_start, block_end, start_col=1, end_col=8):
    for c in range(start_col, end_col+1):
        cell = ws.cell(row=block_end, column=c)
        cell.border = Border(
            left=cell.border.left or thin,
            right=cell.border.right or thin,
            top=cell.border.top or thin,
            bottom=thick
        )
    top_left = ws.cell(row=block_start, column=4)
    top_left.border = Border(
        left=top_left.border.left or thin,
        right=top_left.border.right or thin,
        top=top_left.border.top or thin,
        bottom=thick
    )

def merge_receiver_name(ws, start_row, end_row):
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
        ws.cell(row=start_row, column=4).alignment = Alignment(horizontal='center', vertical='center')

def finalize_orders_sheet(ws):
    min_widths = {
        '플랫폼': 8,
        '주문일시': 16,
        '총 상품결제금액': 14,
        '수취인 이름': 20,
        '상품명 + 옵션명': 50,
        '수량': 10,
        '수취인 전화번호': 16,
        '등록옵션명': 50,
    }
    headers = [cell.value for cell in ws[1]]
    for col in ws.columns:
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ''

        max_len = visual_len(header)
        for cell in col:
            vlen = visual_len(cell.value)
            if vlen > max_len:
                max_len = vlen
            if header == '상품명 + 옵션명' and vlen > 50:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            if header == '등록옵션명':
                cell.number_format = '@'

        auto_width = int(max_len * 0.5)
        if header == '등록옵션명':
            auto_width = int(max_len * 0.5) + 4
        target_width = max(auto_width, min_widths.get(header, 12))
        ws.column_dimensions[col_letter].width = target_width

    for r in range(2, ws.max_row + 1):
        prod_cell = ws.cell(row=r, column=5)
        pclen = visual_len(prod_cell.value)
        if pclen > 40:
            ws.row_dimensions[r].height = 34
        if not (prod_cell.alignment and prod_cell.alignment.wrap_text):
            ws.row_dimensions[r].height = 24

def append_coupang_block(ws, coupang_orders):
    current_row = ws.max_row + 1
    for od in coupang_orders:
        block_start = current_row
        ordered_at = _fmt_dt(od.get('orderedAt') or od.get('orderDate', ''))

        total_price = 0.0
        for item in od.get('orderItems', []):
            price = _to_float(item.get('orderPrice', item.get('price', 0)))
            qty = _to_int(item.get('shippingCount', 1), 1)
            total_price += price * qty
        total_price_str = f"{int(total_price):,}원"

        receiver_name = (
            (od.get('shippingAddress') or {}).get('name', '') or
            (od.get('receiver') or {}).get('name', '')
        )

        item_names = []
        total_qty = 0
        for item in od.get('orderItems', []):
            name = item.get('sellerProductName') or item.get('vendorItemName') or item.get('productName') or ""
            option = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
            qty = _to_int(item.get('shippingCount', 1), 1)
            total_qty += qty
            if name and option and option != name:
                item_names.append(f"{name} / {option}")
            else:
                item_names.append(name or option)
        product_info = " / ".join([x for x in item_names if x])
        total_qty = total_qty or 1

        phone = (
            (od.get('shippingAddress') or {}).get('safeNumber') or
            (od.get('receiver') or {}).get('safeNumber') or
            (od.get('receiver') or {}).get('phone') or
            (od.get('receiver') or {}).get('receiverPhone') or ''
        )

        option_names = []
        for item in od.get('orderItems', []):
            option_name = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
            if option_name:
                option_names.append(str(option_name))
        option_name_str = ", ".join(option_names)

        ws.append([
            "쿠팡", ordered_at, total_price_str, receiver_name,
            product_info, total_qty, phone, option_name_str
        ])
        current_row += 1

        apply_border_block(ws, block_start, current_row - 1, 1, 8)
        merge_receiver_name(ws, block_start, current_row - 1)
        apply_thick_bottom(ws, block_start, current_row - 1, 1, 8)

def append_godo_sets(ws, grouped_orders):
    current_row = ws.max_row + 1
    for grp in grouped_orders:
        block_start = current_row
        first_parent = True

        for s in grp["sets"]:
            p = s["parent"]
            goodsCd  = (p.get('goodsCd') or '').strip()
            goodsNm  = (p.get('goodsNm') or p.get('goodsNmStandard') or '').strip()
            opt_text = (p.get('optionTextInfo') or '').strip()
            qty      = _to_int(p.get('goodsCnt', 1), 1)
            price    = _to_float(p.get('goodsPrice', 0.0), 0.0)

            product_info_parent = f"{goodsCd} / {opt_text}" if opt_text else (goodsNm or goodsCd)
            reg_option_value = goodsCd

            set_total = price * (qty or 1)
            for add in s["children"]:
                add_qty   = _to_int(add.get('goodsCnt', 1), 1)
                add_price = _to_float(add.get('goodsPrice', 0.0), 0.0)
                set_total += add_price * add_qty
            total_price_str = f"{int(set_total):,}원"

            ws.append([
                "고도몰",
                grp["orderedAt"] if first_parent else "",
                total_price_str,
                grp["receiver"]["name"] if first_parent else "",
                product_info_parent,
                (qty or 1),
                grp["receiver"]["phone"] if first_parent else "",
                reg_option_value
            ])
            current_row += 1
            first_parent = False

            prow = current_row - 1
            pcell = ws.cell(row=prow, column=5)
            pcell.font = Font(bold=True)
            pcell.alignment = Alignment(horizontal='left', vertical='center')
            pcell.fill = PatternFill(start_color="FFF7F7F7", end_color="FFF7F7F7", fill_type="solid")

            for add in s["children"]:
                add_name = (add.get('goodsNm') or add.get('goodsNmStandard') or '').strip()
                add_qty  = _to_int(add.get('goodsCnt', 1), 1)
                ws.append(["", "", "", "", f"+ {add_name}", add_qty, "", ""])
                current_row += 1
                crow = current_row - 1
                ccell = ws.cell(row=crow, column=5)
                ccell.font = Font(italic=True, color="00666666")
                ccell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        apply_border_block(ws, block_start, current_row - 1, 1, 8)
        merge_receiver_name(ws, block_start, current_row - 1)
        apply_thick_bottom(ws, block_start, current_row - 1, 1, 8)
