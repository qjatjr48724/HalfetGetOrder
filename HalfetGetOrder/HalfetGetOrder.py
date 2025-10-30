# 1. 라이브러리 불러오기
import os
import time
import hmac, hashlib
import urllib.parse
import urllib.request
import ssl
import requests
import openpyxl
import json
import unicodedata
import math
from datetime import date, datetime, timedelta
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import xmltodict
# from keys import partner_key, godo_key, cp_accesskey, cp_secretkey  # ← 키를 분리해두셨다면 사용

# ─────────────────────────────────────────────────────────
# 공통 사용 변수
# ─────────────────────────────────────────────────────────
content_type = "application/json;charset=UTF-8"

os.environ['TZ'] = 'GMT+0'
datetime_signed = time.strftime('%y%m%d') + 'T' + time.strftime('%H%M%S') + 'Z'
method = "GET"

# 날짜범위 - 시작날짜(오늘날짜 - 7일) / 종료날짜(오늘날짜)
created_from = str(date.today() - timedelta(days=7))
created_to   = str(date.today())
print("조회기간:", created_from, "~", created_to)

# ─────────────────────────────────────────────────────────
# API 키/기본 설정
# ─────────────────────────────────────────────────────────
# 쿠팡 API KEY
cp_accesskey = "7b6058d9-7745-4cf8-9881-cf4a469c0512"
cp_secretkey = "a4e91ae56e91a47b696712dd29008a405e8d4c25"

# 고도몰 API KEY
partner_key = "MCVFRCVENSVCNiUyOCUxMSVEQ28="
godo_key = "JTBBJUNFJTQwZCUxMiVCOCUzQSVBRCVDQW0lRjMlODclRjUlQUElQjIlQ0QlRUYlMDMlQkUlQzklREIlOTh6bU0lMTElNUIlREElM0I0JTFCJUZCcCUxMyVFRCU0MCU3RTUlMUYlRjk="

# ─────────────────────────────────────────────────────────
# 추가옵션 캐시 설정
# ─────────────────────────────────────────────────────────
ADD_GOODS_CACHE = "godo_add_goods_cache.json"
REFRESH_ADD_GOODS = False  # 평소에는 False. 새 옵션 추가 시 True로 1회 실행하여 캐시 갱신.

# ─────────────────────────────────────────────────────────
# 경로(엑셀 저장 위치)
# ─────────────────────────────────────────────────────────
BASE_DIR = r"C:/Users/UserK/Desktop/[작업물]/[판매] 주문수집 프로그램"
os.makedirs(BASE_DIR, exist_ok=True)
today_date = date.today().strftime('%Y%m%d')
WAYBILL_XLSX = os.path.join(BASE_DIR, f"대한통운 송장등록_{today_date}.xlsx")
ORDER_XLSX   = os.path.join(BASE_DIR, f"주문수집_{today_date}.xlsx")

# ─────────────────────────────────────────────────────────
# 공통 유틸/헬퍼
# ─────────────────────────────────────────────────────────
def first_non_empty(*vals):
    for v in vals:
        if v is not None and str(v).strip() != "":
            return v
    return ""

def _to_int(x, default=0):
    try:
        return int(float(str(x).strip()))
    except Exception:
        return default

def _to_float(x, default=0.0):
    try:
        return float(str(x).strip())
    except Exception:
        return default

def _as_list(x):
    if x is None:
        return []
    return x if isinstance(x, list) else [x]

def _fmt_dt(s):
    """문자열 날짜를 'YYYY.MM.DD HH:MM'로 최대한 정규화"""
    if not s:
        return ""
    s = str(s).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y.%m.%d %H:%M")
        except Exception:
            pass
    # yyyy-mm-dd만 온 경우
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%Y.%m.%d 00:00")
    except Exception:
        return s

def get_box_count_from_items(items):
    """3대까지 1박스, 이후 4대마다 1박스 추가"""
    if not items:
        return 1
    try:
        total_qty = sum(int(i.get('quantity', 1)) for i in items)
        if total_qty <= 3:
            return 1
        return 1 + math.ceil((total_qty - 3) / 4)
    except Exception:
        return 1

def visual_len(s):
    """전각(W/F/A)은 2, 그 외는 1로 계산해서 가시 길이 추정"""
    if s is None:
        return 0
    total = 0
    for ch in str(s):
        ea = unicodedata.east_asian_width(ch)
        total += 2 if ea in ('W', 'F', 'A') else 1
    return total

# ─────────────────────────────────────────────────────────
# 추가옵션 캐시 로드/저장
# ─────────────────────────────────────────────────────────
def load_add_goods_map_from_cache(path=ADD_GOODS_CACHE):
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                # {addGoodsNo: {"goodsNm": str, "goodsPrice": float}}
                return {str(k): {"goodsNm": v.get("goodsNm",""),
                                 "goodsPrice": float(v.get("goodsPrice",0.0))}
                        for k, v in data.items()}
    except Exception as e:
        print("추가옵션 캐시 로드 오류:", e)
    return {}

def save_add_goods_map_to_cache(add_map, path=ADD_GOODS_CACHE):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(add_map, f, ensure_ascii=False, indent=2)
        print(f"추가옵션 캐시 저장 완료: {path} (항목 {len(add_map)}개)")
    except Exception as e:
        print("추가옵션 캐시 저장 오류:", e)

# ─────────────────────────────────────────────────────────
# 고도몰 추가상품 목록 조회 (캐시 우선)
# ─────────────────────────────────────────────────────────
def fetch_godomall_add_goods_map():
    add_goods_map = load_add_goods_map_from_cache()
    if REFRESH_ADD_GOODS or not add_goods_map:
        print("추가옵션 캐시가 없거나 갱신 플래그가 True → OpenHub에서 새로 수집합니다.")
        godo_add_url = (
            "https://openhub.godo.co.kr/godomall5/goods/Goods_Add_Search.php?"
            f"partner_key={partner_key}&key={godo_key}"
        )
        r = requests.post(godo_add_url, timeout=30)
        ctype2 = (r.headers.get('Content-Type') or '').lower()
        if 'euc-kr' in ctype2 or 'cp949' in ctype2:
            r.encoding = 'cp949'
        elif not r.encoding:
            r.encoding = 'utf-8'
        add_xml = r.text
        add_json = xmltodict.parse(add_xml)

        def _txt(d, k):
            v = d.get(k)
            return v.strip() if isinstance(v, str) else ("" if v is None else str(v))
        def _price(d):
            for k in ("goodsPrice", "price", "salePrice", "fixedPrice", "consumerPrice"):
                if d.get(k) not in (None, ""):
                    try:
                        return float(str(d[k]).strip())
                    except:
                        pass
            return 0.0

        root_add = add_json.get("data", {}) if isinstance(add_json, dict) else {}
        ret_add  = root_add.get("return", {}) or {}
        cands = None
        for k in ("add_goods_data", "addGoodsData", "goods_data", "goodsData", "list", "rows"):
            if ret_add.get(k) is not None:
                cands = ret_add.get(k)
                break

        items = _as_list(cands)
        if items and isinstance(items[0], dict) and any(subk in items[0] for subk in ("item", "goods", "row")):
            tmp = []
            for it in items:
                for subk in ("item", "goods", "row"):
                    if it.get(subk) is not None:
                        tmp.extend(_as_list(it[subk]))
            if tmp:
                items = tmp

        add_goods_map = {}
        for it in items:
            add_no = _txt(it, "addGoodsNo") or _txt(it, "add_goods_no") or _txt(it, "goodsNo")
            name   = _txt(it, "goodsNm") or _txt(it, "goodsNmStandard") or _txt(it, "goods_name")
            price  = _price(it)
            if add_no and name:
                add_goods_map[add_no] = {"goodsNm": name, "goodsPrice": price}

        print("추가상품 정의 수(신규 수집):", len(add_goods_map))
        save_add_goods_map_to_cache(add_goods_map)
    else:
        print("추가옵션 캐시 사용:", len(add_goods_map), "개")
    return add_goods_map

add_goods_map = fetch_godomall_add_goods_map()

# ─────────────────────────────────────────────────────────
# 고도몰 주문 조회 (XML)
# ─────────────────────────────────────────────────────────
godo_order_url = (
    "https://openhub.godo.co.kr/godomall5/order/Order_Search.php?"
    f"partner_key={partner_key}&key={godo_key}"
    f"&startDate={created_from}&endDate={created_to}"
    "&dateType=order&orderStatus=g1"
)
godo_req = requests.post(godo_order_url, timeout=30)
ctype = (godo_req.headers.get('Content-Type') or '').lower()
if 'euc-kr' in ctype or 'cp949' in ctype:
    godo_req.encoding = 'cp949'
elif not godo_req.encoding:
    godo_req.encoding = 'utf-8'
godo_xml_text = godo_req.text
godo_json = xmltodict.parse(godo_xml_text)

# ─────────────────────────────────────────────────────────
# 쿠팡 api 연동
# ─────────────────────────────────────────────────────────
cp_domain = "https://api-gateway.coupang.com"
cp_venderId = "A01093941"

cp_path = f"/v2/providers/openapi/apis/api/v4/vendors/{cp_venderId}/ordersheets"
cp_query = urllib.parse.urlencode({
    "createdAtFrom": created_from,
    "createdAtTo": created_to,
    "status": "INSTRUCT"
})
message = datetime_signed + method + cp_path + cp_query
signature = hmac.new(cp_secretkey.encode('utf-8'), message.encode('utf-8'), hashlib.sha256).hexdigest()
authorization = (
    f"CEA algorithm=HmacSHA256, access-key={cp_accesskey}, signed-date={datetime_signed}, signature={signature}"
)
cp_url = f"{cp_domain}{cp_path}?{cp_query}"
req = urllib.request.Request(cp_url)
req.add_header("Content-type", content_type)
req.add_header("Authorization", authorization)
req.get_method = lambda: method

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

try:
    resp = urllib.request.urlopen(req, context=ctx)
    coupang_body = resp.read().decode(resp.headers.get_content_charset() or "utf-8")
except Exception as e:
    print("쿠팡 API 호출 오류:", e)
    coupang_body = ""

# ─────────────────────────────────────────────────────────
# 정규화: 쿠팡 / 고도몰
# ─────────────────────────────────────────────────────────
def normalize_coupang_orders(coupang_body):
    try:
        data = json.loads(coupang_body)
    except Exception:
        return []
    orders = data.get('data') or data.get('content') or []
    norm = []
    for od in orders:
        ship = od.get('shippingAddress') or {}
        recv = od.get('receiver') or {}
        orderer = od.get('orderer', {}) or {}
        name = first_non_empty(ship.get('name'), recv.get('name'), orderer.get('name'))
        phone = first_non_empty(
            ship.get('safeNumber'), recv.get('safeNumber'),
            ship.get('phone'), ship.get('phoneNo'),
            recv.get('receiverPhone'), orderer.get('phone')
        )
        addr1 = first_non_empty(ship.get('address1'), recv.get('addr1'))
        addr2 = first_non_empty(ship.get('address2'), recv.get('addr2'))
        zipcode = first_non_empty(ship.get('zipcode'), recv.get('zipCode'))
        items_raw = od.get('orderItems', []) or []
        items = [{"quantity": _to_int(first_non_empty(it.get('shippingCount'), it.get('quantity'), 1), 1)}
                 for it in items_raw]
        norm.append({
            "channel": "coupang",
            "name": name, "phone": phone, "addr1": addr1, "addr2": addr2,
            "zipcode": zipcode, "items": items,
            "raw": od
        })
    return norm

def normalize_godomall_orders_from_xml(json_data):
    try:
        root = json_data.get('data', {})
        ret = root.get('return', {}) or {}
        od_list = _as_list(ret.get('order_data'))
    except Exception:
        return []

    norm = []
    for od in od_list:
        info = od.get('orderInfoData') or {}
        def get_text(d, key):
            v = d.get(key)
            return (v or "").strip() if isinstance(v, str) else (v or "")
        name = get_text(info, 'receiverName')
        safe_fl = (get_text(info, 'receiverUseSafeNumberFl') or '').lower() == 'y'
        safe_no = get_text(info, 'receiverSafeNumber')
        phone = safe_no if (safe_fl and safe_no) else (get_text(info, 'receiverPhone') or get_text(info, 'receiverCellPhone'))
        zipcode = get_text(info, 'receiverZonecode') or get_text(info, 'receiverZipcode')
        addr1 = get_text(info, 'receiverAddress')
        addr2 = get_text(info, 'receiverAddressSub')
        og_list = _as_list(od.get('orderGoodsData'))
        if not addr1 and og_list:
            va = og_list[0].get('visitAddress')
            if isinstance(va, str) and va.strip():
                addr1 = va.strip()
                addr2 = ""
        items = []
        for og in og_list:
            qty_raw = og.get('goodsCnt', 1)
            try:
                qty = int(str(qty_raw).strip() or '1')
            except Exception:
                qty = 1
            items.append({'quantity': qty})
        norm.append({
            "channel": "godomall",
            "name": name, "phone": phone, "addr1": addr1,
            "addr2": addr2, "zipcode": zipcode, "items": items,
            "raw": od
        })
    return norm

# ─────────────────────────────────────────────────────────
# ① 대한통운 송장등록용 엑셀 생성
# ─────────────────────────────────────────────────────────
wb1 = openpyxl.Workbook()
sheet1 = wb1.active
sheet1.title = "판매 주문수집"

first_col1 = [
    '예약구분', '집하예정일', '받는분성명', '받는분전화번호', '받는분기타연락처',
    '받는분우편번호', '받는분주소(전체, 분할)', '운송장번호', '고객주문번호',
    '품목명', '박스수량', '박스타입', '기본운임', '배송메세지1',
    '배송메세지2', '품목명', '운임구분'
]
sheet1.append(first_col1)

header_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
header_font = Font(bold=True)
header_align = Alignment(horizontal="center", vertical="center")
for cell in sheet1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

orders_all = []
if coupang_body:
    orders_all += normalize_coupang_orders(coupang_body)
orders_all += normalize_godomall_orders_from_xml(godo_json)

for od in orders_all:
    name = od.get("name", "")
    phone = od.get("phone", "")
    addr1 = od.get("addr1", "")
    addr2 = od.get("addr2", "")
    zipcode = od.get("zipcode", "")
    address = f"{addr1} {addr2}".strip()
    box_cnt = get_box_count_from_items(od.get("items", []))

    row = [
        "일반",
        date.today().strftime('%Y%m%d'),
        name, phone, "", zipcode, address,
        "", "", "", box_cnt, "", "", "", "", "", ""
    ]
    sheet1.append(row)

# 가운데 정렬 및 열 너비 자동 조정
center_align = Alignment(horizontal="center", vertical="center")
for row in sheet1.iter_rows():
    for cell in row:
        cell.alignment = center_align

for column_cells in sheet1.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in column_cells)
    col_letter = column_cells[0].column_letter
    sheet1.column_dimensions[col_letter].width = max_len * 1.3 + 2

# 저장 (덮어쓰기)
if os.path.exists(WAYBILL_XLSX):
    os.remove(WAYBILL_XLSX)
wb1.save(WAYBILL_XLSX)
print(f"✅ 엑셀 저장 완료: {WAYBILL_XLSX}")

# ─────────────────────────────────────────────────────────
# ② 주문수집(쿠팡 + 고도몰) 엑셀 생성
#   - 고도몰: 본상품=부모행, 추가옵션=하위행
#   - 등록옵션명(고도몰): goodsCd
# ─────────────────────────────────────────────────────────
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = "주문내역"

# 헤더 행 (플랫폼 포함)
header_cols = ['플랫폼', '주문일시', '상품결제금액', '수취인 이름', '상품명 + 옵션명', '수량', '수취인 전화번호', '등록옵션명']
sheet2.append(header_cols)

# 스타일
for cell in sheet2[1]:
    cell.fill = header_fill
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ========== 2-1) 쿠팡 → rows ==========
platform_name = "쿠팡"
if coupang_body:
    try:
        resp_json = json.loads(coupang_body)
        orders = resp_json.get('data') or resp_json.get('content', []) or []

        for od in orders:
            # 주문일시
            ordered_at = _fmt_dt(od.get('orderedAt') or od.get('orderDate', ''))

            # 상품 결제금액 합계
            total_price = 0.0
            for item in od.get('orderItems', []):
                price = _to_float(item.get('orderPrice', item.get('price', 0)))
                qty = _to_int(item.get('shippingCount', 1), 1)
                total_price += price * qty
            total_price_formatted = f"{_to_int(total_price):,}"

            # 수취인 이름
            receiver_name = (
                (od.get('shippingAddress') or {}).get('name', '') or
                (od.get('receiver') or {}).get('name', '')
            )

            # 상품명 + 옵션명
            item_names = []
            for item in od.get('orderItems', []):
                name = item.get('sellerProductName') or item.get('vendorItemName') or item.get('productName') or ""
                option = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
                if name and option and option != name:
                    item_names.append(f"{name} / {option}")
                else:
                    item_names.append(name or option)
            product_info = " / ".join([x for x in item_names if x])

            # 주문 수량
            total_qty = sum(_to_int(item.get('shippingCount', 1), 1) for item in od.get('orderItems', [])) or 1

            # 수취인 전화번호 (안심번호 우선)
            phone = (
                (od.get('shippingAddress') or {}).get('safeNumber') or
                (od.get('receiver') or {}).get('safeNumber') or
                (od.get('receiver') or {}).get('phone') or
                (od.get('receiver') or {}).get('receiverPhone') or ''
            )

            # 등록옵션명 (쿠팡 기준: 옵션명 모음)
            option_names = []
            for item in od.get('orderItems', []):
                option_name = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
                if option_name:
                    option_names.append(str(option_name))
            option_name_str = ", ".join(option_names)

            sheet2.append([
                platform_name,
                ordered_at,
                total_price_formatted,
                receiver_name,
                product_info,
                total_qty,
                phone,
                option_name_str
            ])

    except Exception as e:
        print("⚠️ 쿠팡 JSON 파싱 오류:", e)

# ========== 2-2) 고도몰 → rows (부모행 + 하위행) ==========
platform_name = "고도몰"
try:
    root = godo_json.get('data', {}) if isinstance(godo_json, dict) else {}
    ret = root.get('return', {}) or {}
    od_list = _as_list(ret.get('order_data'))

    for od in od_list:
        ordered_at = _fmt_dt(od.get('orderDate', ''))

        info = od.get('orderInfoData') or {}
        receiver_name = (info.get('receiverName') or "").strip()

        safe_fl = str(info.get('receiverUseSafeNumberFl') or '').strip().lower() == 'y'
        safe_no = (info.get('receiverSafeNumber') or '').strip()
        phone = safe_no if (safe_fl and safe_no) else (
            (info.get('receiverPhone') or info.get('receiverCellPhone') or '').strip()
        )

        # 본상품(부모행 정보)
        parent_lines = []   # "goodsCd / opt_text"
        goods_cd_list = []  # 등록옵션명 칼럼용
        total_qty = 0
        total_price = 0.0

        for og in _as_list(od.get('orderGoodsData')):
            goodsCd  = (og.get('goodsCd') or '').strip()
            goodsNm  = (og.get('goodsNm') or og.get('goodsNmStandard') or '').strip()
            opt_text = (og.get('optionTextInfo') or '').strip()
            qty      = _to_int(og.get('goodsCnt', 1), 1)
            price    = _to_float(og.get('goodsPrice', 0.0), 0.0)

            total_qty   += qty
            total_price += price * qty
            if goodsCd:
                goods_cd_list.append(goodsCd)

            parent_lines.append(f"{goodsNm} / {opt_text}" if opt_text else (goodsNm or goodsCd))

        # 추가옵션(하위행 + 금액 합산)
        add_price_sum = 0.0
        add_display_rows = []  # “+ 옵션명 x수량”

        add_containers = []
        for key in ("orderAddGoodsData", "addGoodsData", "order_add_goods_data", "order_add_goods"):
            val = od.get(key)
            if val is not None:
                add_containers.append(val)

        add_items_all = []
        for cont in add_containers:
            items = _as_list(cont)
            if items and isinstance(items[0], dict) and any(sk in items[0] for sk in ("item", "row", "goods")):
                tmp = []
                for it in items:
                    for sk in ("item", "row", "goods"):
                        if it.get(sk) is not None:
                            tmp.extend(_as_list(it.get(sk)))
                if tmp:
                    items = tmp
            add_items_all.extend(items)

        for add in _as_list(add_items_all):
            add_no   = (add.get("addGoodsNo") or add.get("add_goods_no") or add.get("goodsNo") or "").strip()
            add_qty  = _to_int(add.get("goodsCnt", add.get("addGoodsCnt", 1)), 1)
            add_line_price = _to_float(add.get("goodsPrice", add.get("addGoodsPrice", 0.0)), 0.0)

            # 캐시에서 이름/가격
            add_nm = ""
            add_def_price = 0.0
            if add_no and add_no in add_goods_map:
                add_nm = add_goods_map[add_no]["goodsNm"]
                add_def_price = float(add_goods_map[add_no]["goodsPrice"] or 0.0)

            use_price = add_line_price if add_line_price > 0 else add_def_price
            add_price_sum += use_price * add_qty

            display_name = add_nm or (add.get("goodsNm") or add.get("goodsNmStandard") or "").strip()
            if not display_name and add_no:
                display_name = f"추가상품({add_no})"
            if display_name:
                add_display_rows.append(f"+ {display_name} x{add_qty}")

        # 총액 보정
        if total_price <= 0:
            settle = _to_float(od.get('settlePrice', 0.0), 0.0)
            total_goods_price = _to_float(od.get('totalGoodsPrice', 0.0), 0.0)
            total_price = settle or total_goods_price

        # 추가옵션 금액 합산 → 부모행만 금액에 반영
        total_price += add_price_sum

        total_price_formatted = f"{_to_int(total_price):,}"
        product_info_parent = " / ".join([x for x in parent_lines if x]) or "(본상품 없음)"
        reg_option_value = ", ".join([gc for gc in goods_cd_list if gc])  # 등록옵션명(=goodsCd)

        # 부모행
        sheet2.append([
            platform_name,
            ordered_at,
            total_price_formatted,
            receiver_name,
            product_info_parent,   # goodsCd / opt_text
            (total_qty or 1),
            phone,
            reg_option_value
        ])

        # 하위행(추가옵션들)
        for text in add_display_rows:
            sheet2.append([
                "", "", "", "", text, "", "", ""
            ])

except Exception as e:
    print("⚠️ 고도몰 XML 파싱 오류(옵션 하위행):", e)

# 2-마지막) 시트 서식(가운데 정렬 + 열 너비 자동)
min_widths = {
    '플랫폼': 8,
    '주문일시': 16,
    '상품결제금액': 12,
    '수취인 이름': 20,
    '상품명 + 옵션명': 32,
    '수량': 10,
    '수취인 전화번호': 16,
    '등록옵션명': 50,
}
headers = [cell.value for cell in sheet2[1]]

for col in sheet2.columns:
    col_idx = col[0].column
    col_letter = get_column_letter(col_idx)
    header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ''

    max_len = visual_len(header)
    for cell in col:
        vlen = visual_len(cell.value)
        if vlen > max_len:
            max_len = vlen
        # 기본: 가운데 정렬
        if header == '상품명 + 옵션명' and vlen > 40:
            cell.alignment = center_wrap
        else:
            cell.alignment = center
        # 등록옵션명: 텍스트 서식
        if header == '등록옵션명':
            cell.number_format = '@'

    auto_width = int(max_len * 0.6) + 6
    if header == '등록옵션명':
        auto_width = int(max_len * 0.65) + 8
    target_width = max(auto_width, min_widths.get(header, 12))
    sheet2.column_dimensions[col_letter].width = target_width

# 저장(덮어쓰기)
if os.path.exists(ORDER_XLSX):
    os.remove(ORDER_XLSX)
wb2.save(ORDER_XLSX)
print(f"✅ 엑셀 저장 완료: {ORDER_XLSX}")
