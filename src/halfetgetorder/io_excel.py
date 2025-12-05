import os
import sys
import json
import openpyxl
from datetime import date
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from .utils import visual_len, _to_int, _to_float
from .utils import _fmt_dt, get_box_count_from_items

header_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
thin = Side(style="thin", color="000000")
thick = Side(style="thick", color="000000")

def get_project_root() -> str:
    """
    프로젝트 루트 경로를 반환.

    - 소스에서 실행할 때:
        io_excel.py 기준으로 ../../ 올라간 폴더 (HalfetGetOrder)
    - PyInstaller exe로 실행할 때:
        exe가 위치한 폴더 (dist) 기준
    """
    # PyInstaller로 빌드된 실행 파일 여부
    if getattr(sys, "frozen", False):
        # exe가 있는 폴더
        exe_dir = os.path.dirname(sys.executable)
        return exe_dir

    # 일반 파이썬 실행일 때 (python -m halfetgetorder)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.abspath(os.path.join(base_dir, "..", ".."))


# ─────────────────────────────────────────────────────────
# Rich Text(한 셀 안에 서로 다른 스타일) 지원 여부 체크
# ─────────────────────────────────────────────────────────
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    RICH_TEXT_AVAILABLE = True
except ImportError:
    RICH_TEXT_AVAILABLE = False


def create_orders_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주문내역"
    # C열(총 상품결제금액)과 D열(수취인 이름) 사이에 '체크' 열 추가
    headers = [
        '플랫폼',           # A
        '주문일시',         # B
        '총 상품결제금액',   # C
        '체크',             # D (신규)
        '수취인 이름',      # E
        '상품명 + 옵션명',  # F
        '수량',             # G
        '수취인 전화번호',  # H
        '등록옵션명',       # I
        '배송메세지',       # J
    ]
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
    return wb, ws


def apply_border_block(ws, start_row, end_row, start_col=1, end_col=10):
    for r in range(start_row, end_row+1):
        for c in range(start_col, end_col+1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_thick_bottom(ws, block_start, block_end, start_col=1, end_col=10):
    for c in range(start_col, end_col+1):
        cell = ws.cell(row=block_end, column=c)
        cell.border = Border(
            left=cell.border.left or thin,
            right=cell.border.right or thin,
            top=cell.border.top or thin,
            bottom=thick
        )
    # 굵은 테두리 시작 기준 컬럼도 수취인 이름(E열=5번)로 변경
    top_left = ws.cell(row=block_start, column=5)
    top_left.border = Border(
        left=top_left.border.left or thin,
        right=top_left.border.right or thin,
        top=top_left.border.top or thin,
        bottom=thick
    )


def merge_receiver_name(ws, start_row, end_row):
    # 수취인 이름이 이제 5열(E)이므로 5번 컬럼 기준으로 병합
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)
        ws.cell(row=start_row, column=5).alignment = Alignment(horizontal='center', vertical='center')


def finalize_orders_sheet(ws):
    ws.sheet_view.zoomScale = 75
    min_widths = {
        '플랫폼': 8,
        '주문일시': 16,
        '총 상품결제금액': 14,
        '체크': 6,
        '수취인 이름': 20,
        '상품명 + 옵션명': 70,
        '수량': 10,
        '수취인 전화번호': 16,
        '등록옵션명': 46,
        '배송메세지': 50
    }
    headers = [cell.value for cell in ws[1]]
    for col in ws.columns:
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ''

        max_len = visual_len(header)
        for cell in col:
            vlen = visual_len(cell.value)
            if vlen > max_len:
                max_len = vlen

            # 🔹 상품명 + 옵션명 / 배송메세지 둘 다 긴 경우 줄바꿈 허용
            if header in ('상품명 + 옵션명', '배송메세지') and vlen > 50:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

            if header == '등록옵션명':
                cell.number_format = '@'

        auto_width = int(max_len * 0.5)
        if header == '등록옵션명':
            auto_width = int(max_len * 0.5) + 4
        target_width = max(auto_width, min_widths.get(header, 12))
        ws.column_dimensions[col_letter].width = target_width

    # 상품명+옵션명 열 인덱스: 6열(F)
    # 배송메세지 열 인덱스: 10열(J)
    for r in range(2, ws.max_row + 1):
        prod_cell = ws.cell(row=r, column=6)
        memo_cell = ws.cell(row=r, column=10)

        pclen = visual_len(prod_cell.value)
        mlen = visual_len(memo_cell.value)

        # 두 컬럼 중 더 긴 쪽 기준으로 높이 결정
        base_len = max(pclen, mlen)

        rd = ws.row_dimensions[r]

        # 이미 다른 데서 높이를 지정한 행(예: 부모행 height=65)은 건드리지 않는다
        if rd.height is not None:
            continue

        if base_len > 40:
            rd.height = 34
        else:
            rd.height = 24

    # 🔹 체크 열(D열에 값이 있고, F열이 '+ '로 시작하지 않는 = 부모행만 색상 변경)
    last_row = ws.max_row
    if last_row >= 2:
        fill_checked = PatternFill(
            start_color="FFE6FFCC",
            end_color="FFE6FFCC",
            fill_type="solid"
        )

        # AND(
        #   LEN($D2)>0,          → 체크 열에 뭔가 들어있고
        #   LEFT($F2,2)<>" + "   → F열이 "+ " 로 시작하지 않음 = 자식행이 아님
        # )
        rule = FormulaRule(
            formula=['AND(LEN($D2)>0, LEFT($F2,2)<>" + ")'],
            fill=fill_checked
        )

        # A2 ~ J{마지막 행}까지 적용 → 실제로는 부모행만 색이 들어감
        ws.conditional_formatting.add(f"A2:J{last_row}", rule)


def append_coupang_block(ws, coupang_orders):
    current_row = ws.max_row + 1
    for od in coupang_orders:
        block_start = current_row
        ordered_at = _fmt_dt(od.get('orderedAt') or od.get('orderDate', ''))

        total_price = 0.0
        for item in od.get('orderItems', []):
            price = _to_float(item.get('orderPrice', item.get('price', 0)))
            qty = _to_int(item.get('shippingCount', 1), 1)
            total_price += price * qty
        total_price_str = f"{int(total_price):,}원"

        receiver_name = (
            (od.get('shippingAddress') or {}).get('name', '') or
            (od.get('receiver') or {}).get('name', '')
        )

        item_names = []
        total_qty = 0
        for item in od.get('orderItems', []):
            name = item.get('sellerProductName') or item.get('vendorItemName') or item.get('productName') or ""
            option = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
            qty = _to_int(item.get('shippingCount', 1), 1)
            total_qty += qty
            if name and option and option != name:
                item_names.append(f"{name} / {option}")
            else:
                item_names.append(name or option)
        product_info = " / ".join([x for x in item_names if x])
        total_qty = total_qty or 1

        phone = (
            (od.get('shippingAddress') or {}).get('safeNumber') or
            (od.get('receiver') or {}).get('safeNumber') or
            (od.get('receiver') or {}).get('phone') or
            (od.get('receiver') or {}).get('receiverPhone') or ''
        )

        option_names = []
        for item in od.get('orderItems', []):
            option_name = item.get('sellerProductItemName') or item.get('vendorItemName') or ""
            if option_name:
                option_names.append(str(option_name))
        option_name_str = ", ".join(option_names)

        # 🔹 쿠팡 배송메세지: parcelPrintMessage
        coupang_memo = od.get('parcelPrintMessage', '') or ''

        # A:플랫폼, B:주문일시, C:총금액, D:체크(빈칸), E:수취인, F:상품+옵션, G:수량, H:전화, I:등록옵션명, J:배송메세지
        ws.append([
            "쿠팡",
            ordered_at,
            total_price_str,
            "",                # 체크 열(사용자가 나중에 수동으로 ☑ 등 입력)
            receiver_name,
            product_info,
            total_qty,
            phone,
            option_name_str,
            coupang_memo,
        ])
        current_row += 1

        # 테두리/굵은 라인 범위 1~10열로 확장
        apply_border_block(ws, block_start, current_row - 1, 1, 10)
        merge_receiver_name(ws, block_start, current_row - 1)
        apply_thick_bottom(ws, block_start, current_row - 1, 1, 10)


# ─────────────────────────────────────────────────────────
# 고도몰 추가상품 가져오는 코드(추가상품 json 파일이 없을때만 생성하도록 돌아감)
# ─────────────────────────────────────────────────────────
def load_godo_add_goods_map(path: str | None = None) -> dict:
    """
    godo_add_goods_all.json 로드.

    - 기본 위치: 프로젝트 루트(get_project_root()) 바로 아래.
      (소스 실행시: HalfetGetOrder/, exe 실행시: HalfetGetOrder.exe 가 있는 폴더)

    - 파일이 없으면:
        * 개발(소스) 실행: build_godo_add_goods_all.main() 을 한 번 호출하여 자동 생성 시도
        * exe 실행(PyInstaller): 자동 생성하지 않고, 경고만 출력 후 빈 dict 반환
    """
    project_root = get_project_root()
    default_path = os.path.join(project_root, "godo_add_goods_all.json")

    # 1) 인자로 path가 들어온 경우 처리
    if path:
        # 상대경로면 프로젝트 루트 기준으로
        if not os.path.isabs(path):
            candidate = os.path.join(project_root, path)
        else:
            candidate = path

        # 실제 파일이 있으면 그걸 사용
        if os.path.exists(candidate):
            resolved_path = candidate
        else:
            # 없으면 기본 경로로
            resolved_path = default_path
    else:
        resolved_path = default_path

    # 2) 최종 경로에 파일이 없을 때
    if not os.path.exists(resolved_path):
        # exe 환경에서는 자동 생성 X
        if getattr(sys, "frozen", False):
            print(
                "⚠️ godo_add_goods_all.json 파일을 찾을 수 없습니다.\n"
                "   exe 환경에서는 자동 생성하지 않고, "
                "추가옵션 매핑 없이 계속 진행합니다."
            )
            return {}

        # 개발(소스) 환경일 때만 자동 생성
        print("⚠️ godo_add_goods_all.json 이 없어 처음 한 번 생성합니다...")
        try:
            from . import build_godo_add_goods_all
        except ImportError:
            print(
                "⚠️ build_godo_add_goods_all 모듈을 찾을 수 없습니다. "
                "추가옵션 매핑 없이 계속 진행합니다."
            )
            return {}

        try:
            build_godo_add_goods_all.main()
        except Exception as e:
            print(f"⚠️ godo_add_goods_all.json 생성 중 오류: {e}")
            return {}

        # main()이 default_path에 저장했을 가능성이 높으므로 다시 확인
        if (not os.path.exists(resolved_path)) and os.path.exists(default_path):
            resolved_path = default_path

        if not os.path.exists(resolved_path):
            print(
                "⚠️ godo_add_goods_all.json 을 생성했지만, 파일을 찾지 못했습니다.\n"
                "   추가옵션 매핑 없이 계속 진행합니다."
            )
            return {}

    # 3) 최종 경로에서 로드
    try:
        with open(resolved_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"⚠️ godo_add_goods_all.json 로드 중 오류: {e}")
        return {}

    return data
    
def _parse_short_desc_to_specs(short_desc: str) -> tuple[str, str]:
    """
    shortDescription 예시:
      'DeLL Latitude 5501 / Intel® Core™ i7-9850H / NVIDIA GeForce MX150 / NVMe SSD 512G / DDR4 32G / FHD ... / 윈도우11'

    - '/' 로 나눈 뒤
      index 3 → SSD 파트 (예: 'NVMe SSD 512G')
      index 4 → RAM 파트 (예: 'DDR4 32G')
    - 각 파트를 마지막 토큰만 쓰지 않고, **있는 그대로** 반환한다.
    """
    if not short_desc:
        return "", ""

    parts = [p.strip() for p in str(short_desc).split("/") if p.strip()]

    # SSD: 3번째(인덱스 3)
    ssd = parts[3].strip() if len(parts) > 3 else ""
    # RAM: 4번째(인덱스 4)
    ram = parts[4].strip() if len(parts) > 4 else ""

    # (RAM, SSD) 순서로 반환
    return ram, ssd



def _build_base_specs_from_raw(raw) -> dict:
    """
    raw 를 {상품번호: {ram, ssd}} 형태로 정규화.
    지원 형태:
      1) 딕셔너리:
         {
           "1000001": { "ram": "16G", "ssd": "512G" }
           "1000002": { "shortDescription": "..." }
           "1000003": "DeLL Latitude 5501 / ... / NVMe SSD 512G / DDR4 32G / ..."
         }

      2) 리스트:
         [
           { "goodsNo": "1000001", "ram": "16G", "ssd": "512G" },
           { "goodsNo": "1000002", "shortDescription": "..." },
           { "goodsCd": "NB-5501", "shortDescription": "..." },
           ...
         ]
    """
    base_specs: dict[str, dict[str, str]] = {}

    # case 1: dict
    if isinstance(raw, dict):
        for key, val in raw.items():
            goods_key = str(key).strip()
            if not goods_key:
                continue

            ram = ""
            ssd = ""

            if isinstance(val, dict):
                ram = str(val.get("ram", "")).strip()
                ssd = str(val.get("ssd", "")).strip()
                short_desc = str(val.get("shortDescription", "")).strip()

                # ram/ssd 없으면 shortDescription에서 뽑기
                if short_desc and (not ram or not ssd):
                    ram2, ssd2 = _parse_short_desc_to_specs(short_desc)
                    ram = ram or ram2
                    ssd = ssd or ssd2
            else:
                # 값이 그냥 shortDescription 문자열인 경우
                short_desc = str(val).strip()
                if short_desc:
                    ram, ssd = _parse_short_desc_to_specs(short_desc)

            base_specs[goods_key] = {"ram": ram, "ssd": ssd}

    # case 2: list
    elif isinstance(raw, list):
        for row in raw:
            if not isinstance(row, dict):
                continue

            goods_key = str(
                row.get("goodsNo") or row.get("goodsCd") or ""
            ).strip()
            if not goods_key:
                continue

            ram = str(row.get("ram", "")).strip()
            ssd = str(row.get("ssd", "")).strip()
            short_desc = str(row.get("shortDescription", "")).strip()

            if short_desc and (not ram or not ssd):
                ram2, ssd2 = _parse_short_desc_to_specs(short_desc)
                ram = ram or ram2
                ssd = ssd or ssd2

            base_specs[goods_key] = {"ram": ram, "ssd": ssd}

    return base_specs


def load_godo_base_specs_map(path: str | None = None) -> dict:
    """
    고도몰 상품 기본 RAM/SSD 사양 로드.

    우선순위:
      1) 인자로 받은 path
      2) 프로젝트 루트의 godo_base_specs.json
      3) 프로젝트 루트의 godo_goods_all.json (goods_search 결과 전체)
    """
    project_root = get_project_root()

    candidates: list[str] = []
    if path:
        candidates.append(path)
    candidates.append(os.path.join(project_root, "godo_base_specs.json"))
    candidates.append(os.path.join(project_root, "godo_goods_all.json"))

    for p in candidates:
        if not p:
            continue
        if not os.path.exists(p):
            continue

        try:
            with open(p, "r", encoding="utf-8") as f:
                raw = json.load(f)
            specs = _build_base_specs_from_raw(raw)
        except Exception as e:
            print(f"⚠️ 기본 사양 파일({p})을 읽는 중 오류: {e}")
            continue

        if specs:
            print(f"[라벨] 고도몰 기본 RAM/SSD 사양 {len(specs)}건 로드 ({p})")
            return specs

    print("⚠️ godo_base_specs.json / godo_goods_all.json 을 찾지 못했습니다. 고도몰 라벨의 RAM/SSD는 비워둡니다.")
    return {}


def get_godo_base_ram_ssd(parent: dict, base_specs_map: dict) -> tuple[str, str]:
    """
    고도몰 parent(본상품) 한 건에 대해 기본 RAM/SSD 를 조회.
    - 우선 goodsNo 로 찾고
    - 없으면 goodsCd 로도 한 번 더 찾아본다.
    """
    goods_no = str(parent.get("goodsNo") or "").strip()
    goods_cd = str(parent.get("goodsCd") or "").strip()

    spec = None
    if goods_no:
        spec = base_specs_map.get(goods_no)
    if spec is None and goods_cd:
        spec = base_specs_map.get(goods_cd)

    if not spec:
        return "", ""

    ram = str(spec.get("ram", "")).strip()
    ssd = str(spec.get("ssd", "")).strip()
    return ram, ssd
    

def load_godo_goods_map(path: str | None = None) -> dict:
    """
    goods_search로 미리 만들어둔 godo_goods_all.json 로드.
    key: goodsNo
    value: goods_search 응답 전체(dict)
    """
    if path is None:
        project_root = get_project_root()
        path = os.path.join(project_root, "godo_goods_all.json")

    if not os.path.exists(path):
        print("⚠️ godo_goods_all.json 파일을 찾을 수 없습니다. 기본 RAM/SSD는 비워둡니다.")
        return {}

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)
    

# ─────────────────────────────────────────────────────────
# shortDescription에서 기본 RAM/SSD 뽑는 함수 추가
# ─────────────────────────────────────────────────────────
def get_base_specs_from_short_description(parent: dict, goods_map: dict) -> tuple[str, str]:
    """
    - 우선 parent(주문의 본상품) 안에 shortDescription 이 있으면 그걸 쓰고,
    - 없으면 godo_goods_all.json(goods_map)에서 goodsNo 로 찾아서 shortDescription을 가져온다.

    shortDescription 예시:
      DeLL Latitude 5501 / Intel® Core™ i7-9850H / NVIDIA GeForce MX150 /
      NVMe SSD 512G / DDR4 32G / FHD 1920×1080 해상도 (15.6인치) / 윈도우11

    / 로 split 한 후:
      0: 모델명
      1: CPU
      2: 그래픽
      3: SSD
      4: RAM
      5: 해상도
      6: 윈도우 버전

    여기서
      - 기본 SSD  → parts[3]
      - 기본 RAM  → parts[4]
    를 **그대로** 사용한다.
    """
    # 1) 주문 데이터에 바로 shortDescription 이 들어있으면 우선 사용
    short_desc = (parent.get("shortDescription") or "").strip()

    # 2) 없으면 goodsNo로 godo_goods_all.json 에서 찾아본다
    goods_no = str(parent.get("goodsNo") or "").strip()
    if not short_desc and goods_no and goods_map:
        if isinstance(goods_map, dict):
            goods_info = goods_map.get(goods_no)
            if isinstance(goods_info, dict):
                short_desc = (goods_info.get("shortDescription") or
                              goods_info.get("short_desc") or "").strip()
        elif isinstance(goods_map, list):
            # goods_map 이 리스트인 경우 (goods_search 결과를 그대로 저장한 형태)
            for row in goods_map:
                if not isinstance(row, dict):
                    continue
                key = str(row.get("goodsNo") or row.get("goodsCd") or "").strip()
                if key == goods_no:
                    short_desc = (row.get("shortDescription") or
                                  row.get("short_desc") or "").strip()
                    break

    if not short_desc:
        return "", ""

    parts = [p.strip() for p in short_desc.split("/")]

    # 최소한 SSD(3), RAM(4) 까지는 있어야 한다
    if len(parts) <= 4:
        return "", ""

    ssd_part = parts[3].strip()
    ram_part = parts[4].strip()

    # (RAM, SSD) 순서대로 반환
    return ram_part, ssd_part





def extract_specs_from_godo_children_using_map(children: list, add_goods_map: dict):
    """
    고도몰 '추가상품(children)' 리스트와 godo_add_goods_all.json을 사용해
    RAM / SSD / 옵션 문자열을 추출.

    godo_add_goods_all.json 구조:
    {
      "1000000015": { "name": "고급 노트북 가방 구매", "summary": "OPT:가방" },
      "1000000078": { "name": "용량 256G→NVMe SSD 1TB로 UP↑", "summary": "SSD:1TB" },
      ...
    }
    """
    ram = None
    ssd = None
    options: list[str] = []
    missing_ids = set()

    for add in children:
        add_no = str(add.get("addGoodsNo") or "").strip()
        if not add_no:
            continue

        entry = add_goods_map.get(add_no)
        if not entry:
            # 매핑표에 없는 추가옵션 번호
            missing_ids.add(add_no)
            continue

        summary = (entry.get("summary") or "").strip()
        if not summary:
            # summary(요약이름 B)를 아직 안 채운 경우
            missing_ids.add(add_no)
            continue

        # prefix 기반 파싱: "RAM:16G", "SSD:1TB", "OPT:원키" ...
        prefix, sep, value = summary.partition(":")
        prefix = prefix.strip().upper()
        value = value.strip() if sep else summary  # 콜론 없으면 전체를 value로

        if prefix == "RAM" and value:
            ram = value
        elif prefix == "SSD" and value:
            ssd = value
        else:
            # OPT:..., 혹은 prefix 없는 경우 모두 옵션으로 취급
            if value:
                options.append(value)

    if missing_ids:
        print(f"[라벨] 매핑되지 않은 추가옵션 번호: {', '.join(sorted(missing_ids))}")

    # 옵션 중복 제거 + 정렬
    options = sorted(set(options))
    option_str = " / ".join(options) if options else ""

    return ram or "", ssd or "", option_str



# ─────────────────────────────────────────────────────────
# 쿠팡 추가상품 가져오는 코드
# ─────────────────────────────────────────────────────────
def extract_specs_from_coupang_item(item: dict, keyskin_models: list[str] | None = None):
    """
    쿠팡 orderItems[*]에서 RAM / SSD / 옵션 추출.
    - RAM: sellerProductItemName.split()[3]
    - SSD: sellerProductItemName.split()[2]
    - 옵션: [리브레, 원키] + (모델명에 키워드 포함되면 키스킨)
    """
    seller_item_name = item.get("sellerProductItemName") or ""
    tokens = seller_item_name.split()

    ram = ""
    ssd = ""

    if len(tokens) > 3:
        ram = tokens[3]
    if len(tokens) > 2:
        ssd = tokens[2]

    # 옵션 기본값
    options = ["리브레", "원키"]

    # 모델명 기반 키스킨 추가
    if keyskin_models:
        model_name = (
            item.get("sellerProductName")
            or item.get("vendorItemName")
            or item.get("productName")
            or ""
        )
        for kw in keyskin_models:
            if kw and kw in model_name:
                options.append("키스킨")
                break

    option_str = " / ".join(options)
    return ram, ssd, option_str


def append_godo_sets(ws, grouped_orders):
    """
    고도몰 주문을 엑셀 주문내역 시트에 추가.
    - 부모행(본상품)의 '상품명 + 옵션명' 셀(6열)에:
        상품명
        orderoptionInfo
      이렇게 줄바꿈해서 표시.
    - 상품명은 볼드,
      orderoptionInfo는 회색+기울임(가능하면).
    """
    current_row = ws.max_row + 1
    for grp in grouped_orders:
        block_start = current_row
        first_parent = True

        for s in grp["sets"]:
            p = s["parent"]
            goodsCd  = (p.get('goodsCd') or '').strip()
            goodsNm  = (p.get('goodsNm') or p.get('goodsNmStandard') or '').strip()
            opt_text = (p.get('optionTextInfo') or '').strip()
            qty      = _to_int(p.get('goodsCnt', 1), 1)
            price    = _to_float(p.get('goodsPrice', 0.0), 0.0)

            # ▶ 상품명 + 옵션명(부모셀) 구성 로직
            product_name = goodsNm or goodsCd

            # 1) 먼저 orderoptionInfo / orderOptionInfo 에 사람이 읽기 좋게 들어있는지 확인
            option_info = (p.get('orderoptionInfo') or p.get('orderOptionInfo') or '').strip()

            # 2) 없으면 optionInfo(JSON 문자열)를 파싱해서 "옵션명: 값" 형태로 뽑기
            if not option_info:
                raw_opt = (p.get('optionInfo') or '').strip()
                if raw_opt:
                    try:
                        opt_list = json.loads(raw_opt)  # [[옵션명, 옵션값, ...], [...], ...]
                        parts = []
                        for opt in opt_list:
                            if isinstance(opt, (list, tuple)) and len(opt) >= 2:
                                name = str(opt[0]).strip()
                                val = str(opt[1]).strip()
                                if name and val:
                                    parts.append(f"{name}: {val}")
                        option_info = "\n".join(parts)
                    except Exception:
                        option_info = ""

            if option_info:
                product_info_parent = f"{product_name}\n{option_info}"
            else:
                product_info_parent = product_name

            reg_option_value = goodsCd

            set_total = price * (qty or 1)
            for add in s["children"]:
                add_qty   = _to_int(add.get('goodsCnt', 1), 1)
                add_price = _to_float(add.get('goodsPrice', 0.0), 0.0)
                set_total += add_price * add_qty
            total_price_str = f"{int(set_total):,}원"

            order_memo = grp.get("orderMemo", "") or grp.get("orderInfo", {}).get("orderMemo", "")

            # A:플랫폼, B:주문일시, C:총금액, D:체크, E:수취인, F:상품+옵션, G:수량, H:전화, I:등록옵션명, J:배송메세지
            ws.append([
                "고도몰",
                grp["orderedAt"] if first_parent else "",
                total_price_str,
                "",   # 체크 열(사용자 수동 입력용)
                grp["receiver"]["name"] if first_parent else "",
                product_info_parent,
                (qty or 1),
                grp["receiver"]["phone"] if first_parent else "",
                reg_option_value,
                order_memo if first_parent else ""
            ])
            current_row += 1
            first_parent = False

            # 부모 셀 스타일링 (상품명+옵션명: 6열)
            prow = current_row - 1
            pcell = ws.cell(row=prow, column=6)

            if option_info and RICH_TEXT_AVAILABLE:
                pcell.value = CellRichText(
                    TextBlock(
                        text=product_name,
                        font=InlineFont(
                            b=True
                        )
                    ),
                    TextBlock(
                        text="\n" + option_info,
                        font=InlineFont(
                            i=True,
                            color="00666666"
                        )
                    ),
                )
            else:
                pcell.value = product_info_parent
                pcell.font = Font(bold=True)

            pcell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            pcell.fill = PatternFill(start_color="FFF7F7F7", end_color="FFF7F7F7", fill_type="solid")

            ws.row_dimensions[prow].height = 65

            # 자식(추가옵션) 행
            for add in s["children"]:
                add_name = (add.get('goodsNm') or add.get('goodsNmStandard') or '').strip()
                add_qty  = _to_int(add.get('goodsCnt', 1), 1)
                # A~J 열 구조에 맞춰서 한 칸씩 밀어줌
                ws.append(["", "", "", "", "", f"+ {add_name}", add_qty, "", "", ""])
                current_row += 1
                crow = current_row - 1
                ccell = ws.cell(row=crow, column=6)
                ccell.font = Font(italic=True, color="00666666")
                ccell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        apply_border_block(ws, block_start, current_row - 1, 1, 10)
        merge_receiver_name(ws, block_start, current_row - 1)
        apply_thick_bottom(ws, block_start, current_row - 1, 1, 10)





# ─────────────────────────────────────────────────────────
# 라벨 출력 전용 엑셀파일 만드는 코드
# ─────────────────────────────────────────────────────────
def create_label_workbook(
    coupang_orders: list,
    godo_grouped_orders: list,
    godo_add_goods_map_path: str | None = None,
):
    """
    라벨 출력용 엑셀 워크북 생성.

    헤더:
      플랫폼 / 이름 / 모델명 / 램 / SSD / 옵션

    - coupang_orders: 쿠팡 원본 주문 리스트
    - godo_grouped_orders: 고도몰 grouped_orders 리스트
    """
    # 고도몰 추가상품 매핑 로드
    try:
        add_goods_map = load_godo_add_goods_map(godo_add_goods_map_path)
    except FileNotFoundError:
        print("⚠️ godo_add_goods_all.json 파일을 찾을 수 없습니다. (고도몰 라벨에는 추가옵션 매핑이 반영되지 않습니다.)")
        add_goods_map = {}

    # 🔹 고도몰 기본 RAM/SSD 사양 로드 (shortDescription 기반)
    base_specs_map = load_godo_base_specs_map()
    missing_base_spec_ids: set[str] = set()

    # 🔹 shortDescription fallback 용 전체 상품 정보 (goods_search 결과)
    try:
        godo_goods_map = load_godo_goods_map()
    except Exception as e:
        print(f"⚠️ godo_goods_all.json 로드 중 오류: {e}")
        godo_goods_map = {}

    # 쿠팡 키스킨 모델 리스트 (원하면 json으로 분리해도 됨)
    keyskin_models = [
        "그램 17",
        "Latitude 5520",
        "키스킨 포함",
        "키보드 키스킨",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "라벨"

    # 헤더
    headers = ["플랫폼", "이름", "모델명", "램", "SSD", "옵션"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(
            start_color="D8E4BC", end_color="D8E4BC", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ─────────────────────────────────────────
    # 1) 쿠팡 라벨
    # ─────────────────────────────────────────
    for od in coupang_orders:
        receiver_name = (
            (od.get("shippingAddress") or {}).get("name", "")
            or (od.get("receiver") or {}).get("name", "")
        )

        for item in od.get("orderItems", []):
            model_name = (
                item.get("sellerProductName")
                or item.get("vendorItemName")
                or item.get("productName")
                or ""
            )

            ram, ssd, option_str = extract_specs_from_coupang_item(
                item,
                keyskin_models=keyskin_models,
            )

            # 👉 shippingCount(수량) 만큼 같은 행을 반복해서 추가
            qty = _to_int(item.get("shippingCount", 1), 1)
            if qty <= 0:
                qty = 1

            for _ in range(qty):
                ws.append(
                    [
                        "쿠",           # 플랫폼
                        receiver_name, # 이름
                        model_name,    # 모델명
                        ram,
                        ssd,
                        option_str,    # 옵션
                    ]
                )

    # ─────────────────────────────────────────
    # 2) 고도몰 라벨
    #   - 자사몰 주문은 역순(최근 주문이 아래로)
    #   - goodsCnt(수량) 만큼 행 반복
    #   - 모델명 셀: "모델명\noptionInfo"
    # ─────────────────────────────────────────
    for grp in reversed(godo_grouped_orders or []):
        receiver_name = grp.get("receiver", {}).get("name", "")

        for s in grp.get("sets", []):
            parent = s.get("parent", {})
            children = s.get("children", []) or []

            model_name = (parent.get("goodsCd") or "").strip()

            # 1) 기본 RAM/SSD: 우선 base_specs_map 사용
            base_ram, base_ssd = get_godo_base_ram_ssd(parent, base_specs_map)

            # 1-1) 부족하면 shortDescription 을 직접 파싱해서 보완
            if (not base_ram or not base_ssd) and godo_goods_map:
                try:
                    ram2, ssd2 = get_base_specs_from_short_description(
                        parent, godo_goods_map
                    )
                    base_ram = base_ram or ram2
                    base_ssd = base_ssd or ssd2
                except Exception:
                    pass

            if not (base_ram or base_ssd):
                goods_no = str(parent.get("goodsNo") or "").strip()
                key = goods_no or model_name
                if key:
                    missing_base_spec_ids.add(key)

            # 2) 부모 상품의 optionInfo 문자열 만들기
            #    - orderoptionInfo / orderOptionInfo 에 사람이 읽기 좋은 포맷이 있으면 그걸 우선 사용
            #    - 없으면 optionInfo(JSON) 파싱해서 "옵션명: 옵션값 / ..." 형태로 생성
            option_info = (
                (parent.get("orderoptionInfo") or "").strip()
                or (parent.get("orderOptionInfo") or "").strip()
            )

            if not option_info:
                raw_opt = (parent.get("optionInfo") or "").strip()
                if raw_opt:
                    try:
                        opt_list = json.loads(raw_opt)  # [[옵션명, 옵션값, ...], ...]
                        parts: list[str] = []
                        for opt in opt_list:
                            if isinstance(opt, (list, tuple)) and len(opt) >= 2:
                                name = str(opt[0]).strip()
                                val = str(opt[1]).strip()
                                if name and val:
                                    parts.append(f"{name}: {val}")
                        option_info = " / ".join(parts)
                    except Exception:
                        option_info = ""

            # 3) 모델명 셀 값: "모델명" 또는 "모델명\noptionInfo"
            model_cell_value = model_name
            if option_info:
                model_cell_value = f"{model_name}\n{option_info}"

            # 4) 추가옵션(가방/원키/복구 등)은 옵션열(F)로
            _, _, option_str = extract_specs_from_godo_children_using_map(
                children, add_goods_map
            )

            # 5) 본상품 수량(goodsCnt) 만큼 행을 반복해서 추가
            qty = _to_int(parent.get("goodsCnt", 1), 1)
            if qty <= 0:
                qty = 1

            for _ in range(qty):
                ws.append(
                    [
                        "자",               # 플랫폼(자사몰)
                        receiver_name,      # 이름
                        model_cell_value,   # 모델명 + optionInfo(줄바꿈)
                        base_ram,           # 램
                        base_ssd,           # SSD
                        option_str,         # 옵션(추가상품 요약)
                    ]
                )

    # 정렬 & 열 너비 세팅
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter == "C":  # 모델명 열은 줄바꿈 허용
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    width_map = {
        "A": 10,  # 플랫폼
        "B": 18,  # 이름
        "C": 45,  # 모델명(+옵션)
        "D": 12,  # 램
        "E": 12,  # SSD
        "F": 30,  # 옵션
    }
    for col, w in width_map.items():
        ws.column_dimensions[col].width = w

    ws.sheet_view.zoomScale = 90

    # 기본 사양 맵은 있는데도 매칭이 안 된 상품들 로그
    if base_specs_map and missing_base_spec_ids:
        print(
            "[라벨] RAM/SSD 기본사양을 찾지 못한 고도몰 상품번호/코드: "
            + ", ".join(sorted(missing_base_spec_ids))
        )

    return wb, ws




def create_waybill_workbook(coupang_orders):
    """
    대한통운 송장등록용 엑셀 워크북 생성.
    - 시트명: '판매 주문수집'
    - 열 구조: 기존 단일 파일 버전의 first_col1 과 동일
    - coupang_orders: coupang.normalize_coupang_orders(...) 결과 리스트
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판매 주문수집"
    ws.sheet_view.zoomScale = 75
    header = [
        '예약구분', '집하예정일', '받는분성명', '받는분전화번호', '받는분기타연락처',
        '받는분우편번호', '받는분주소(전체, 분할)', '운송장번호', '고객주문번호',
        '품목명', '박스수량', '박스타입', '기본운임', '배송메세지1',
        '배송메세지2', '품목명', '운임구분'
    ]
    ws.append(header)

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    today_str = date.today().strftime('%Y%m%d')

    for od in coupang_orders:
        name = od.get("name", "")
        phone = od.get("phone", "")
        addr1 = od.get("addr1", "")
        addr2 = od.get("addr2", "")
        zipcode = od.get("zipcode", "")
        address = f"{addr1} {addr2}".strip()

        box_cnt = get_box_count_from_items(od.get("items", []))

        platform_name = "쿠팡"

        row = [
            "일반",
            today_str,
            name,
            phone,
            "",
            zipcode,
            address,
            "",
            "",
            "",
            box_cnt,
            "",
            "",
            "",
            platform_name,
            "",
            ""
        ]
        ws.append(row)

    center_align = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align

    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max_len * 1.3 + 2

    return wb, ws
